#!/usr/bin/env python3
"""Audit workbook formulas, lineage, policies, and cached checks against a spec."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import range_boundaries


FORMULA_ERRORS = {
    "#CALC!",
    "#DIV/0!",
    "#FIELD!",
    "#NAME?",
    "#N/A",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#SPILL!",
    "#VALUE!",
}
CELL_REFERENCE = re.compile(
    r"(?:(?:'((?:[^']|'')+)'|([A-Za-z_][A-Za-z0-9_. ]*))!)?"
    r"\$?([A-Za-z]{1,3})\$?([0-9]+)"
)


def split_ref(reference: str) -> tuple[str, str]:
    if "!" not in reference:
        raise ValueError(f"Cell reference must include a sheet: {reference}")
    sheet, cell = reference.rsplit("!", 1)
    return sheet.strip("'"), cell.replace("$", "").upper()


def canonical_ref(reference: str) -> str:
    sheet, cell = split_ref(reference)
    return f"{sheet.casefold()}!{cell.upper()}"


def formula_references(formula: str, current_sheet: str) -> set[str]:
    references: set[str] = set()
    for match in CELL_REFERENCE.finditer(formula):
        quoted_sheet, plain_sheet, column, row = match.groups()
        sheet = quoted_sheet.replace("''", "'") if quoted_sheet else plain_sheet
        sheet = sheet or current_sheet
        references.add(f"{sheet.casefold()}!{column.upper()}{row}")
    return references


def direct_formula_reference(formula: str, current_sheet: str) -> str | None:
    expression = formula.strip()
    if not expression.startswith("="):
        return None
    expression = expression[1:].strip()
    if expression.startswith("+"):
        expression = expression[1:].strip()
    match = CELL_REFERENCE.fullmatch(expression)
    if not match:
        return None
    quoted_sheet, plain_sheet, column, row = match.groups()
    sheet = quoted_sheet.replace("''", "'") if quoted_sheet else plain_sheet
    sheet = sheet or current_sheet
    return f"{sheet.casefold()}!{column.upper()}{row}"


def iter_range(workbook: Any, reference: str) -> list[Any]:
    sheet_name, coordinates = split_ref(reference)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Missing sheet {sheet_name}")
    min_col, min_row, max_col, max_row = range_boundaries(coordinates)
    sheet = workbook[sheet_name]
    return [
        sheet.cell(row=row, column=column)
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    ]


def cell(workbook: Any, reference: str) -> Any:
    sheet_name, coordinate = split_ref(reference)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Missing sheet {sheet_name}")
    return workbook[sheet_name][coordinate]


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def values_equal(actual: Any, expected: Any, tolerance: float = 0.0) -> bool:
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        if not (math.isfinite(float(actual)) and math.isfinite(float(expected))):
            return actual == expected
        return abs(float(actual) - float(expected)) <= tolerance
    return str(actual).strip().casefold() == str(expected).strip().casefold()


def audit(workbook_path: Path, specification: dict[str, Any]) -> dict[str, Any]:
    formulas = load_workbook(workbook_path, data_only=False, read_only=False, keep_links=True)
    cached = load_workbook(workbook_path, data_only=True, read_only=False, keep_links=True)
    failures: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    def fail(control: str, message: str, **details: Any) -> None:
        failures.append({"control": control, "message": message, **details})

    required_sheets = specification.get("required_sheets", [])
    for sheet_name in required_sheets:
        if sheet_name not in formulas.sheetnames:
            fail("required_sheets", f"Missing required sheet: {sheet_name}")

    formula_count = 0
    for sheet in formulas.worksheets:
        for row in sheet.iter_rows():
            for formula_cell in row:
                if is_formula(formula_cell.value):
                    formula_count += 1
                    upper_formula = formula_cell.value.upper()
                    for error in FORMULA_ERRORS:
                        if error in upper_formula:
                            fail(
                                "formula_errors",
                                f"Formula contains {error} at {sheet.title}!{formula_cell.coordinate}",
                            )
    minimum = int(specification.get("min_formula_count", 0))
    if formula_count < minimum:
        fail(
            "min_formula_count",
            f"Workbook contains {formula_count} formulas; expected at least {minimum}",
        )

    external_links = len(getattr(formulas, "_external_links", []))
    if external_links and not specification.get("allow_external_links", False):
        fail("external_links", f"Workbook contains {external_links} external link(s)")

    for sheet in cached.worksheets:
        for row in sheet.iter_rows():
            for cached_cell in row:
                if isinstance(cached_cell.value, str) and cached_cell.value.upper() in FORMULA_ERRORS:
                    fail(
                        "formula_errors",
                        f"Formula error {cached_cell.value} at {sheet.title}!{cached_cell.coordinate}",
                    )

    for reference in specification.get("input_cells", []):
        try:
            value = cell(formulas, reference).value
            if value in (None, ""):
                fail("input_cells", f"Required input is blank: {reference}")
            elif is_formula(value):
                fail("input_cells", f"Authoritative input cannot be a formula: {reference}")
        except (KeyError, ValueError) as error:
            fail("input_cells", str(error), cell=reference)

    for reference in specification.get("formula_cells", []):
        try:
            if not is_formula(cell(formulas, reference).value):
                fail("formula_cells", f"Expected a formula at {reference}")
        except (KeyError, ValueError) as error:
            fail("formula_cells", str(error), cell=reference)

    for policy in specification.get("policy_inputs", []):
        reference = policy["cell"]
        label = policy.get("label", reference)
        try:
            value = cell(formulas, reference).value
            if value in (None, ""):
                fail("policy_inputs", f"Policy input is blank: {label} ({reference})")
                continue
            if is_formula(value):
                fail("policy_inputs", f"Policy input must be editable: {label} ({reference})")
                continue
            if "allowed" in policy and not any(
                values_equal(value, allowed, float(policy.get("tolerance", 0.0)))
                for allowed in policy["allowed"]
            ):
                fail(
                    "policy_inputs",
                    f"Unexpected policy value at {reference}: {value!r}",
                    allowed=policy["allowed"],
                )
        except (KeyError, ValueError) as error:
            fail("policy_inputs", str(error), cell=reference)

    for link in specification.get("authoritative_links", []):
        target = link["target"]
        source = canonical_ref(link["source"])
        try:
            target_cell = cell(formulas, target)
            if not is_formula(target_cell.value):
                fail("authoritative_links", f"Linked output is hardcoded: {target}")
                continue
            references = formula_references(target_cell.value, target_cell.parent.title)
            if source not in references:
                fail(
                    "authoritative_links",
                    f"{target} does not reference authoritative source {link['source']}",
                    formula=target_cell.value,
                )
            elif link.get("direct_only", True):
                direct = direct_formula_reference(target_cell.value, target_cell.parent.title)
                if direct != source:
                    fail(
                        "authoritative_links",
                        f"{target} must be a direct link to {link['source']}",
                        formula=target_cell.value,
                    )
        except (KeyError, ValueError) as error:
            fail("authoritative_links", str(error), cell=target)

    for requirement in specification.get("formula_requirements", []):
        references: list[str] = list(requirement.get("cells", []))
        if "range" in requirement:
            try:
                references.extend(
                    f"{entry.parent.title}!{entry.coordinate}"
                    for entry in iter_range(formulas, requirement["range"])
                )
            except (KeyError, ValueError) as error:
                fail("formula_requirements", str(error), range=requirement["range"])
                continue
        must_reference = {
            canonical_ref(reference) for reference in requirement.get("must_reference_all", [])
        }
        patterns = [re.compile(pattern) for pattern in requirement.get("forbid_regex", [])]
        for reference in references:
            try:
                formula_cell = cell(formulas, reference)
                formula = formula_cell.value
                if not is_formula(formula):
                    fail("formula_requirements", f"Expected a formula at {reference}")
                    continue
                actual_references = formula_references(formula, formula_cell.parent.title)
                missing = sorted(must_reference.difference(actual_references))
                if missing:
                    fail(
                        "formula_requirements",
                        f"{reference} is missing required reference(s)",
                        missing=missing,
                        formula=formula,
                    )
                for pattern in patterns:
                    if pattern.search(formula):
                        fail(
                            "formula_requirements",
                            f"{reference} contains forbidden formula pattern {pattern.pattern!r}",
                            formula=formula,
                        )
            except (KeyError, ValueError) as error:
                fail("formula_requirements", str(error), cell=reference)

    for series in specification.get("copy_series", []):
        try:
            excluded = {canonical_ref(ref) for ref in series.get("exceptions", [])}
            cells = [
                entry
                for entry in iter_range(formulas, series["range"])
                if canonical_ref(f"{entry.parent.title}!{entry.coordinate}") not in excluded
            ]
            if not cells:
                warnings.append(
                    {"control": "copy_series", "message": f"No cells tested: {series['range']}"}
                )
                continue
            origin = cells[0].coordinate
            normalized: list[tuple[str, str]] = []
            for formula_cell in cells:
                formula = formula_cell.value
                if not is_formula(formula):
                    fail(
                        "copy_series",
                        f"Expected a copy-across formula at {formula_cell.parent.title}!{formula_cell.coordinate}",
                    )
                    continue
                try:
                    translated = Translator(formula, origin=formula_cell.coordinate).translate_formula(origin)
                except Exception as error:  # openpyxl raises several formula-specific exceptions
                    fail(
                        "copy_series",
                        f"Could not normalize formula at {formula_cell.parent.title}!{formula_cell.coordinate}: {error}",
                    )
                    continue
                normalized.append((formula_cell.coordinate, translated))
            if normalized:
                expected = normalized[0][1]
                for coordinate, translated in normalized[1:]:
                    if translated != expected:
                        fail(
                            "copy_series",
                            f"Formula family mismatch in {series.get('name', series['range'])}",
                            cell=f"{cells[0].parent.title}!{coordinate}",
                            expected_normalized=expected,
                            actual_normalized=translated,
                        )
        except (KeyError, ValueError) as error:
            fail("copy_series", str(error), range=series.get("range"))

    for status in specification.get("status_cells", []):
        reference = status["cell"]
        try:
            actual = cell(cached, reference).value
            expected = status["equals"]
            if actual is None:
                fail("status_cells", f"Missing cached result at {reference}; recalculate the workbook")
            elif not values_equal(actual, expected, float(status.get("tolerance", 0.0))):
                fail(
                    "status_cells",
                    f"Unexpected status at {reference}: {actual!r}",
                    expected=expected,
                )
        except (KeyError, ValueError) as error:
            fail("status_cells", str(error), cell=reference)

    for check in specification.get("numeric_checks", []):
        try:
            left = cell(cached, check["left"]).value
            right = check.get("expected")
            if "right" in check:
                right = cell(cached, check["right"]).value
            tolerance = float(check.get("tolerance", 0.0))
            if not isinstance(left, (int, float)) or not isinstance(right, (int, float)):
                fail(
                    "numeric_checks",
                    f"Numeric check lacks cached numeric values: {check}",
                    left_value=left,
                    right_value=right,
                )
            elif abs(float(left) - float(right)) > tolerance:
                fail(
                    "numeric_checks",
                    f"Numeric check failed: {check['left']}",
                    left_value=left,
                    right_value=right,
                    tolerance=tolerance,
                )
        except (KeyError, ValueError) as error:
            fail("numeric_checks", str(error), check=check)

    return {
        "workbook": str(workbook_path.resolve()),
        "passed": not failures,
        "formula_count": formula_count,
        "external_link_count": external_links,
        "failure_count": len(failures),
        "warning_count": len(warnings),
        "failures": failures,
        "warnings": warnings,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Audit a financial workbook against its pre-authored model specification."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--spec", required=True, type=Path)
    args = parser.parse_args()

    if not args.workbook.is_file():
        raise FileNotFoundError(args.workbook)
    if not args.spec.is_file():
        raise FileNotFoundError(args.spec)
    with args.spec.open(encoding="utf-8") as handle:
        specification = json.load(handle)
    report = audit(args.workbook, specification)
    print(json.dumps(report, indent=2, default=str))
    if not report["passed"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
